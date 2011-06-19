#!/usr/bin/env python
import datetime
from openpyxl.reader.excel import load_workbook

class SuuntoEntry(object):
    def __init__(self, time, hr, altitude):
        self.time = time
        self.hr = hr
        self.altitude = altitude

class SuuntoXLSXParser(object):
    entries = []
    start_time = 0
    date = 0
    interval = 0
    duration = 0
    max_hr = 0
    avg_hr = 0
    hr_sum = 0
    kcal_sum = 0
    distance = 0
    comment = ''

    def num_to_column(self, n):
        first_dim = n / 26
        second_dim = n % 26

        if first_dim == 0:
            return chr(ord('A')+second_dim)
        else:
            return "%s%s" % (chr(ord('A')-1+first_dim), chr(ord('A')+second_dim))

    def get_laps_samplecols(self, ws):
        laps = 0
        hrcol = 0
        altcol = 0
        for i in range(0, ws.get_highest_column()):
            column = '%s' % self.num_to_column(i)
            coordinate = '%s1' % column
            v =  ws.cell(coordinate=coordinate).value
            if type(v) == str or type(v) == unicode:
                if v.startswith('Move mark'):
                    laps += 1
                elif v.startswith('Move samples'):
                    hrcol = column
                    altcol = self.num_to_column(i+5)
                    break
        return (laps, hrcol, altcol)

    def parse_uploaded_file(self, f):
        wb = load_workbook(filename=f.name)
        ws = wb.get_active_sheet()

        start_dt = datetime.datetime.strptime(ws.cell(coordinate='B3').value, '%Y-%m-%d %H:%M:%S')
        self.kcal_sum = ws.cell(coordinate='D3').value
        self.interval = ws.cell(coordinate='O3').value
        self.duration = ws.cell(coordinate='C3').value
        self.distance = ws.cell(coordinate='E3').value
        self.date = start_dt.date()
        self.start_time = start_dt.time()
        self.comment = '%s %s' % (ws.cell(coordinate='A3').value, ws.cell(coordinate='Z3').value)


        laps, hrcol, altitudecol = self.get_laps_samplecols(ws)

        i = 3
        hr = ws.cell(coordinate='%s%d' % (hrcol, i)).value
        alt = ws.cell(coordinate='%s%d' % (altitudecol, i)).value
        self.max_hr = hr

        while hr != None and alt != None:
            timestamp = start_dt + datetime.timedelta(0, self.interval * (i-3))
            se = SuuntoEntry(timestamp, hr, alt)
            self.entries.append(se)
            self.hr_sum += hr
            if hr > self.max_hr:
                self.max_hr = hr
            i += 1
            hr = ws.cell(coordinate='%s%d' % (hrcol, i)).value
            alt = ws.cell(coordinate='%s%d' % (altitudecol, i)).value
        self.avg_hr = self.hr_sum/(i-1)


if __name__ == '__main__':
    import sys

    p = SuuntoXLSXParser()
    p.parse_uploaded_file(file(sys.argv[1]))

    for entry in p.entries:
        print 'hr: %d altitude: %d' % (entry.hr, entry.altitude)
    print '--------------'
    print 'Date: %s Comment: %s' % (p.date, p.comment)
    print 'Start: %s Duration %s seconds' % (p.start_time, p.duration)
    print 'Distance: %s Sample interval: %s' % (p.distance, p.interval)
    print 'kcal: %s Avg hr: %s' % (p.kcal_sum, p.avg_hr)


